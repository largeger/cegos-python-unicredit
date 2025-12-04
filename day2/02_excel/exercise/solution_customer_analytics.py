import pandas as pd
import glob
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

file_path = 'day2/02_excel/exercise/'

# ---- Verkaufsdateien zusammenführen ----
verkaeufe_files = glob.glob(file_path + 'verkaeufe_*.xlsx')
verkaeufe_all = pd.DataFrame()
for file in verkaeufe_files:
    df = pd.read_excel(file, engine='openpyxl')
    verkaeufe_all = pd.concat([verkaeufe_all, df], ignore_index=True)
print(verkaeufe_all)

# ---- Kundendateien zusammenführen ----
kunden_files = glob.glob(file_path + 'kunden_*.xlsx')
kunden_all = pd.DataFrame()
for file in kunden_files:
    df = pd.read_excel(file, engine='openpyxl')
    kunden_all = pd.concat([kunden_all, df], ignore_index=True)
print (kunden_all)

# ---- Zusammenführen auf Bestell-ID ----
daten = pd.merge(verkaeufe_all, kunden_all, on='Bestell-ID', how='inner')

# ---- Umsatz berechnen ----
daten['Umsatz'] = daten['Preis'] * daten['Menge']

# ---- Top 3 Kunden nach Umsatz ----
top3 = daten.groupby('Kunde')['Umsatz'].sum().sort_values(ascending=False).head(3)
print('Top 3 Kunden nach Umsatz:')
print(top3, '\n')

# ---- Gesamtumsatz pro Stadt ----
umsatz_stadt = daten.groupby('Stadt')['Umsatz'].sum()
print('Gesamtumsatz pro Stadt:')
print(umsatz_stadt, '\n')

# ---- Bonus: Treuebonus ----
daten['Treuebonus'] = daten['Treueprogramm'].apply(lambda x: 0.05 if x == 'Ja' else 0.0)
daten['Umsatz_bereinigt'] = daten['Umsatz'] * (1 - daten['Treuebonus'])
print('Daten mit bereinigtem Umsatz:')
print(daten)

# ---- In Excel speichern und stylen ----
output_file = 'zusammengefuehrte_daten.xlsx'
daten.to_excel(output_file, index=False, engine='openpyxl')


# Workbook laden für Styling
wb = load_workbook(output_file)
ws = wb.active


# Header fett machen und Hintergrundfarbe
header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'),
top=Side(border_style='thin'), bottom=Side(border_style='thin'))


for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = border


# Alle Zellen mit Rahmen versehen
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.border = border


wb.save(output_file)
print(f'Daten erfolgreich in "{output_file}" gespeichert und gestylt.')